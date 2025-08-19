import datetime
import os
from PyQt6.QtWidgets import (
    QWidget,
    QFormLayout,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QTableView,
    QHeaderView,
    QPushButton,
    QTableWidget,
    QAbstractItemView,
    QStyle,
    QDialog,
    QSpinBox,
    QTableWidgetItem,
    QMessageBox,
    QFileDialog,
)
from openpyxl import load_workbook
from PyQt6.QtGui import QFont, QIntValidator, QGuiApplication
from PyQt6 import QtCore
from calendar import monthrange
from utils.vietnamese_currency import convert_number_to_vietnamese
from controllers.finance_controller import FinanceController
from models import IncomeOutcome, Setting
from settings import BASE_DIR


headers = [
    'ID',
    "Ngày",
    "Người nhận",
    "Nội dung",
    "Số tiền",
]


class Form(QDialog):

    def __init__(self, parent=None, month=1, year=datetime.date.today().year, edit=False, id=None):
        self.month = month
        self.year = year
        self.updated = False
        self.id = id

        super(Form, self).__init__(parent)
        form_layout = QFormLayout()
        
        name_label = QLabel("Họ và tên người nhận tiền")
        self.name_input = QLineEdit()
        form_layout.addRow(name_label, self.name_input)

        address_label = QLabel("Địa chỉ")
        self.address_input = QLineEdit()
        form_layout.addRow(address_label, self.address_input)

        reason_label = QLabel("Lý do chi")
        self.reason_input = QLineEdit()
        form_layout.addRow(reason_label, self.reason_input)
        
        amount_label = QLabel("Số tiền")
        validator = QIntValidator(0, 2147483647)
        self.amount_input = QLineEdit()
        self.amount_input.setValidator(validator)
        form_layout.addRow(amount_label, self.amount_input)
        
        amount_text_label = QLabel("Viết bằng chữ")
        self.amount_text_label = QLabel("")
        form_layout.addRow(amount_text_label, self.amount_text_label)
        
        description_label = QLabel("Kèm theo")
        self.description_input = QLineEdit()
        form_layout.addRow(description_label, self.description_input)

        date_label = QLabel("Ngày")
        self.date_input = QSpinBox()
        self.date_input.setMinimum(1)
        self.date_input.setMaximum(monthrange(self.year, self.month)[1])

        today = datetime.date.today()
        if today.month == month:
            self.date_input.setValue(today.day)

        form_layout.addRow(date_label, self.date_input)

        submit_button = QPushButton("Lưu")
    
        cancel_button = QPushButton("Hủy")
        cancel_button.clicked.connect(self.close)
        
        export_button = QPushButton("Xuất file")
        export_button.clicked.connect(self.export_file)
        
        button_layout = QHBoxLayout()
        button_layout.addWidget(submit_button)
        button_layout.addWidget(export_button)
        button_layout.addWidget(cancel_button)
        form_layout.addRow(button_layout)

        self.setLayout(form_layout)

        if not edit:
            submit_button.clicked.connect(self.insert)
            self.setWindowTitle("Thêm chi")
        else:
            submit_button.clicked.connect(self.update)
            self.setWindowTitle("Sửa")
            
        self.amount_input.textChanged.connect(self.format_amount_input)     
            
    def format_amount_input(self, text):
        # Remove all non-digit characters
        cleaned = ''.join(filter(str.isdigit, text))

        if cleaned:
            number = int(cleaned)
            formatted = f"{number:,}"

            # Temporarily block signals to avoid infinite recursion
            self.amount_input.blockSignals(True)
            self.amount_input.setText(formatted)
            self.amount_input.blockSignals(False)

            # Adjust cursor position
            self.amount_input.setCursorPosition(len(formatted))

            # Update labels
            self.amount_text_label.setText(convert_number_to_vietnamese(number))
        else:
            self.amount_input.blockSignals(True)
            self.amount_input.setText("")
            self.amount_input.blockSignals(False)
            self.amount_text_label.setText("")


    def validated(self):
        conditions = [
            self.name_input.text().strip() != '',
            self.address_input.text().strip() != '',
            self.amount_input.text().strip() != '',
            self.amount_input.text().replace(",", "").isdigit(),
        ]

        return all(conditions)

    
    def insert(self):
        if not self.validated():
            return

        data = {
            'type': IncomeOutcome.OUTCOME,
            'name': self.name_input.text(),
            'address': self.address_input.text(),
            'reason': self.reason_input.text(),
            'amount': int(self.amount_input.text().replace(",", "")),
            'description': self.description_input.text(),
            'date': datetime.date(self.year, self.month, self.date_input.value()),
        }

        FinanceController.add_entry(data)
        self.updated = True
        self.accept()

    def update(self):
        if not self.validated():
            return
        
        data = {
            "name": self.name_input.text().strip(),
            "address": self.address_input.text().strip(),
            "reason": self.reason_input.text().strip(),
            "amount": int(self.amount_input.text().replace(",", "").strip()),
            "description": self.description_input.text().strip(),
            "date": datetime.date(self.year, self.month, self.date_input.value()),
        }
        
        result = FinanceController.update_entry(self.id, data)

        if result:
            self.updated = True
            self.accept()
        else:
            QMessageBox.warning(self, "Update Failed", "Could not update the record.")
    
    def export_file(self):
        if not self.validated():
            QMessageBox.warning(self, "Export Failed", "Vui lòng điền đầy đủ thông tin trước khi xuất file.")
            return
        
        template_path = os.path.join(BASE_DIR, "templates", "PHIEU_CHI.xlsx")

        workbook = load_workbook(template_path)
        sheet = workbook.active
        
        values = {
            "{USER_NAME}": self.name_input.text(),
            "{ADDRESS}": self.address_input.text(),
            "{REASON}": self.reason_input.text(),
            "{AMOUNT}": f"{int(self.amount_input.text().replace(',', '')):,}",
            "{AMOUNT_TEXT}": self.amount_text_label.text(),
            "{DESCRIPTION}": self.description_input.text(),
            "{DEPARTMENT}": Setting.get_value("department", ""),
            "{AGENCY}": Setting.get_value("agency", ""),
            "{CHIEF_LABEL}": Setting.get_value("chief_label", "Thủ trưởng"),
            "{CHIEF_ACCOUNTANT_LABEL}": Setting.get_value("chief_accountant_label", "Kế toán trưởng"),
            "{TREASURER_LABEL}": Setting.get_value("treasurer_label", "Thủ quỹ"),
            "{CHIEF_NAME}": Setting.get_value("chief", ""),
            "{CHIEF_ACCOUNTANT_NAME}": Setting.get_value("chief_accountant", ""),
            "{TREASURER_NAME}": Setting.get_value("treasurer", ""),
        }

        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in values:
                    cell.value = values[cell.value]

        save_path, _ = QFileDialog.getSaveFileName(self, "Lưu phiếu thu", "PhieuChi.xlsx", "Excel Files (*.xlsx)")
        if save_path:
            workbook.save(save_path)
            QMessageBox.information(self, "Thành công", f"Phiếu thu đã được lưu tại:\n{save_path}")


class OutcomeView(QWidget):

    def __init__(self, month, year):
        super().__init__()
        self.month = month
        self.year = year
        self.initUI()
        self.loadData()

    def initUI(self):
        screen = QGuiApplication.primaryScreen()
        
        geometry = screen.availableGeometry()
        desktop_width = geometry.width()
        desktop_height = geometry.height()

        width = int(desktop_width)
        height = int(desktop_height)

        self.resize(width, height)

        self.showMaximized()

        vbox1 = QVBoxLayout()

        label = QLabel(f'Tháng {self.month}, Năm {self.year}')
        font = QFont('Times', 15)
        label.setFont(font)
        label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)

        vbox1.addWidget(label)
        vbox1.addSpacing(15)

        hbox1 = QHBoxLayout()
        back_button = QPushButton()
        back_button.setIcon(
            self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowBack)
        )
        back_button.clicked.connect(self.backButtonClicked)

        hbox1.addWidget(back_button)

        hbox1.addStretch()
        
        self.delete_button = QPushButton("Xóa")
        self.delete_button.setVisible(False)
        self.delete_button.clicked.connect(self.deleteSelected)
        hbox1.addWidget(self.delete_button)

        add_button = QPushButton("Thêm")
        add_button.clicked.connect(self.addButtonClicked)

        hbox1.addWidget(add_button)
        vbox1.addLayout(hbox1)
        vbox1.addSpacing(15)

        self.table = self.createTableView()

        vbox1.addWidget(self.table)

        self.setLayout(vbox1)
        self.setWindowTitle("Mục chi")
        self.show()

    def createTableView(self):
        table = QTableWidget()
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        table.setRowCount(10)
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setColumnHidden(0, True)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        table.cellDoubleClicked.connect(self.cellClicked)
        
        table.itemSelectionChanged.connect(self.onSelectionChanged)
        
        return table

    
    def onSelectionChanged(self):
        selected = self.table.selectionModel().selectedRows()
        self.delete_button.setVisible(len(selected) == 1)

    
    def createTableSummary(self):
        table = QTableWidget()
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setRowCount(1)
        table.setColumnCount(3)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        table.setMaximumHeight(70)
        return table


    def loadData(self):
        entries = FinanceController.get_entries(
            entry_type=IncomeOutcome.OUTCOME,
            month=self.month,
            year=self.year
        )
        
        self.table.setRowCount(len(entries))
        for row, entry in enumerate(entries):
            self.table.setItem(row, 0, QTableWidgetItem(str(entry.id)))
            self.table.setItem(row, 1, QTableWidgetItem(entry.date.strftime("%d/%m/%Y")))
            self.table.setItem(row, 2, QTableWidgetItem(entry.name))
            self.table.setItem(row, 3, QTableWidgetItem(entry.reason))
            self.table.setItem(row, 4, QTableWidgetItem(f"{entry.amount:,}"))


    def backButtonClicked(self):
        from .home import Home
        self.screen = Home()
        self.close()


    def addButtonClicked(self):
        self.form = Form(month=self.month, year=self.year)
        self.form.exec()
        
        if self.form.updated:
            self.loadData()


    def cellClicked(self, row):
        item = self.table.item(row, 0)

        try:
            record_id = int(item.text())
        except:
            return
        
        record = FinanceController.get_entry(record_id)
        if not record:
            return
        
        self.form = Form(month=self.month, year=self.year, edit=True, id=record_id)
        self.form.name_input.setText(record.name)
        self.form.address_input.setText(record.address or "")
        self.form.reason_input.setText(record.reason or "")
        self.form.amount_input.setText(f"{record.amount:,}")
        self.form.description_input.setText(record.description or "")
        self.form.date_input.setValue(record.date.day)
        self.form.amount_text_label.setText(convert_number_to_vietnamese(record.amount))

        self.form.exec()

        if self.form.updated:
            self.loadData()
            
    
    def deleteSelected(self):
        row = self.table.currentRow()
        if row < 0:
            return

        id_item = self.table.item(row, 0)
        if id_item is None or not id_item.text().strip().isdigit():
            QMessageBox.warning(self, "Lỗi", "Không xác định được bản ghi để xoá.")
            return

        record_id = int(id_item.text())
        
        date = self.table.item(row, 1).text() if self.table.item(row, 1) else ""
        note = self.table.item(row, 3).text() if self.table.item(row, 3) else ""

        reply = QMessageBox.question(
            self,
            "Xác nhận xoá",
            f"Bạn có chắc muốn xoá bản ghi của “{date}: {note}”?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if reply == QMessageBox.StandardButton.Yes:
            try:
                ok = FinanceController.delete_entry(record_id)
            except Exception as e:
                ok = False

            if ok:
                self.loadData()
                self.delete_button.setVisible(False)  # Ẩn lại sau khi xoá
                QMessageBox.information(self, "Thành công", "Đã xoá bản ghi.")
            else:
                QMessageBox.warning(self, "Xoá thất bại", "Không thể xoá bản ghi.")
