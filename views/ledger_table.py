import calendar
import datetime
import os
from copy import copy
from PyQt6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QHBoxLayout,
    QPushButton,
    QStyle,
    QFileDialog,
    QMessageBox,
)
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from models import IncomeOutcome, Setting
from PyQt6.QtCore import Qt
from settings import BASE_DIR

def copy_row_style(sheet, src_row_idx, dest_row_idx):
    """Copy style từ 1 row chuẩn sang row mới."""
    for col in range(1, sheet.max_column + 1):
        src_cell = sheet.cell(row=src_row_idx, column=col)
        dest_cell = sheet.cell(row=dest_row_idx, column=col)
        if src_cell.has_style:
            dest_cell.font = copy(src_cell.font)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.number_format = copy(src_cell.number_format)
            dest_cell.protection = copy(src_cell.protection)
            dest_cell.alignment = copy(src_cell.alignment)

def unmerge_row_with_style(sheet, row_idx):
    """Unmerge cells in a row but preserve border/style."""
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row <= row_idx <= merged.max_row:
            # Lưu style từ ô gốc
            master_cell = sheet.cell(row=merged.min_row, column=merged.min_col)
            style_copy = {
                "font": copy(master_cell.font),
                "border": copy(master_cell.border),
                "fill": copy(master_cell.fill),
                "number_format": copy(master_cell.number_format),
                "protection": copy(master_cell.protection),
                "alignment": copy(master_cell.alignment),
            }

            # Bỏ merge
            sheet.unmerge_cells(str(merged))

            # Copy style sang toàn bộ vùng merge cũ
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    cell = sheet.cell(row=r, column=c)
                    cell.font = style_copy["font"]
                    cell.border = style_copy["border"]
                    cell.fill = style_copy["fill"]
                    cell.number_format = style_copy["number_format"]
                    cell.protection = style_copy["protection"]
                    cell.alignment = style_copy["alignment"]

class LedgerTableView(QWidget):
    def __init__(self, month, year):
        super().__init__()
        self.month = month
        self.year = year
        self.initUI()
        self.loadData()
    
    def initUI(self):
        self.setWindowTitle("Sổ Thu Chi")
        
        layout = QVBoxLayout()
        
        hbox1 = QHBoxLayout()
        back_button = QPushButton()
        back_button.setIcon(
            self.style().standardIcon(QStyle.StandardPixmap.SP_ArrowBack)
        )
        back_button.clicked.connect(self.backButtonClicked)
        
        export_button = QPushButton("Xuất file")
        export_button.clicked.connect(self.export_file)

        hbox1.addWidget(back_button)
        hbox1.addStretch()
        hbox1.addWidget(export_button)

        hbox1.addStretch()
        
        layout.addLayout(hbox1)
        
        self.table = self.createTableView()
        
        layout.addWidget(self.table)
        
        self.setLayout(layout)
        
        self.showMaximized()
        
    def createTableView(self):
        days_in_month = calendar.monthrange(self.year, self.month)[1]
        
        table = QTableWidget(days_in_month + 2, 7)   # 2 header rows + 10 data rows, 9 columns
        
        table.setSpan(0, 0, 2, 1)
        table.setSpan(0, 1, 1, 2) # "Số phiếu" spans
        table.setSpan(0, 3, 2, 1) # "DIỄN GIẢI" spans
        table.setSpan(0, 4, 1, 3) # "SỐ TIỀN" spans

        # Fill header rows manually
        headers_row1 = ["Ngày tháng", "Số phiếu", "", "DIỄN GIẢI", "SỐ TIỀN", "", ""]
        headers_row2 = ["", "Thu", "Chi", "NỘI DUNG CHI", "THU", "CHI", "TỒN",]
        
        for col, text in enumerate(headers_row1):
            item = QTableWidgetItem(text)
            item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            table.setItem(0, col, item)

        for col, text in enumerate(headers_row2):
            item = QTableWidgetItem(text)
            item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            table.setItem(1, col, item)

        # Optional: center-align header cells
        for row in [0, 1]:
            for col in range(9):
                item = table.item(row, col)
                if item:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        # # Freeze header rows (make them uneditable)
        for col in range(7):
            table.item(0, col).setFlags(Qt.ItemFlag.ItemIsEnabled)
            table.item(1, col).setFlags(Qt.ItemFlag.ItemIsEnabled)
        
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        # Make the table editable from row 2 onward (row 0, 1 are headers)
        table.setVerticalHeaderLabels([""]*(days_in_month + 2))  # hide row numbers
        table.horizontalHeader().hide()
        
        return table
    
    def loadData(self):
        start_date = datetime.date(self.year, self.month, 1)
        end_date = datetime.date(self.year, self.month, calendar.monthrange(self.year, self.month)[1])

        entries = (
            IncomeOutcome.select()
            .where((IncomeOutcome.date >= start_date) & (IncomeOutcome.date <= end_date))
            .order_by(IncomeOutcome.date, IncomeOutcome.created_at)
        )

        self.table.setRowCount(len(entries) + 2)  # 2 dòng header + số phiếu
        self.table.setVerticalHeaderLabels([""]*(len(entries) + 2))

        running_balance = 0
        row_index = 2  # sau header

        for entry in entries:
            # Cập nhật số dư
            if entry.type == IncomeOutcome.INCOME:
                running_balance += entry.amount
            else:
                running_balance -= entry.amount

            # Cột 0: Ngày
            self.table.setItem(row_index, 0, QTableWidgetItem(entry.date.strftime("%d/%m/%Y")))

            # Phiếu thu
            if entry.type == IncomeOutcome.INCOME:
                self.table.setItem(row_index, 1, QTableWidgetItem("1"))
                self.table.setItem(row_index, 4, QTableWidgetItem(f"{entry.amount:,}"))
            # Phiếu chi
            else:
                self.table.setItem(row_index, 2, QTableWidgetItem("1"))
                self.table.setItem(row_index, 5, QTableWidgetItem(f"{entry.amount:,}"))

            # Cột 3: Diễn giải
            self.table.setItem(row_index, 3, QTableWidgetItem(entry.reason or ""))

            # Cột 6: Tồn
            self.table.setItem(row_index, 6, QTableWidgetItem(f"{running_balance:,}"))

            # Căn giữa các cột số liệu
            for col in [0,1,2,4,5,6]:
                item = self.table.item(row_index, col)
                if item:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

            row_index += 1
    
    def backButtonClicked(self):
        from .home import Home
        self.screen = Home()
        self.close()
        
    def export_file(self):
        start_date = datetime.date(self.year, self.month, 1)
        end_date = datetime.date(self.year, self.month, calendar.monthrange(self.year, self.month)[1])

        entries = (
            IncomeOutcome.select()
            .where((IncomeOutcome.date >= start_date) & (IncomeOutcome.date <= end_date))
            .order_by(IncomeOutcome.date, IncomeOutcome.created_at)
        )

        template_path = os.path.join(BASE_DIR, "templates", "SO_QUY_TIEN_MAT.xlsx")
        workbook = load_workbook(template_path)
        sheet = workbook.active

        today = datetime.date.today()
        # Replace placeholders in header
        values = {
            "{DEPARTMENT}": Setting.get_value("department", ""),
            "{AGENCY}": Setting.get_value("agency", ""),
            "{MONTH}/{YEAR}": f"{self.month}/{self.year}",
            "{CURRENT_YEAR}": f", ngày {today.day} tháng {today.month} năm {today.year}",
            "{CHIEF_LABEL}": Setting.get_value("chief_label", "Thủ trưởng"),
            "{CHIEF_ACCOUNTANT_LABEL}": Setting.get_value("chief_accountant_label", "Kế toán trưởng"),
            "{TREASURER_LABEL}": Setting.get_value("treasurer_label", "Thủ quỹ"),
            "{CHIEF_NAME}": Setting.get_value("chief", ""),
            "{CHIEF_ACCOUNTANT_NAME}": Setting.get_value("chief_accountant", ""),
            "{TREASURER_NAME}": Setting.get_value("treasurer", ""),
        }

        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in values:
                    cell.value = values[cell.value.strip()]

        # Find the row index of "TỔNG CỘNG"
        total_row_idx = None
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row[0] and "TỔNG CỘNG" in str(row[0]):
                total_row_idx = row_idx
                break

        if not total_row_idx:
            QMessageBox.warning(self, "Xuất thất bại", "Không tìm thấy dòng TỔNG CỘNG trong template.")
            return

        running_balance = 0
        total_income = 0
        total_outcome = 0

        current_row = total_row_idx - 1  # bắt đầu ngay trước "TỔNG CỘNG"
        template_row_idx = total_row_idx - 2  # row mẫu để copy style
        
        for entry in entries:
            sheet.insert_rows(current_row)
            
            # bỏ merge ở dòng mới để không bị dính gộp
            unmerge_row_with_style(sheet, current_row)
            
            # Copy style từ row mẫu sang row mới
            copy_row_style(sheet, template_row_idx, current_row)

            # Điền dữ liệu
            sheet.cell(row=current_row, column=1, value=entry.date.strftime("%d/%m/%Y"))
            if entry.type == IncomeOutcome.INCOME:
                sheet.cell(row=current_row, column=2, value="1")
                sheet.cell(row=current_row, column=5, value=entry.amount)
                total_income += entry.amount
                running_balance += entry.amount
            else:
                sheet.cell(row=current_row, column=3, value="1")
                sheet.cell(row=current_row, column=6, value=entry.amount)
                total_outcome += entry.amount
                running_balance -= entry.amount

            sheet.cell(row=current_row, column=4, value=entry.reason or "")
            sheet.cell(row=current_row, column=7, value=running_balance)

            current_row += 1

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row[0] and "TỔNG CỘNG" in str(row[0]):
                total_row_idx = row_idx
                break

        sheet.delete_rows(total_row_idx - 1, 1)
        sheet.cell(row=total_row_idx - 1, column=5, value=total_income)
        sheet.cell(row=total_row_idx - 1, column=6, value=total_outcome)
        sheet.cell(row=total_row_idx - 1, column=7, value=running_balance)
        sheet.merge_cells(
            start_row=total_row_idx - 2,
            start_column=1,
            end_row=total_row_idx - 2,
            end_column=3
        )
        cell = sheet.cell(row=total_row_idx - 1, column=1)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.delete_rows(8, 1)
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).strip() == values["{CURRENT_YEAR}"]:
                    col = cell.column
                    row_idx = cell.row
                    if col + 2 <= sheet.max_column:  # tránh vượt giới hạn
                        sheet.merge_cells(start_row=row_idx, start_column=col,
                                          end_row=row_idx, end_column=col+3)
                        merged_cell = sheet.cell(row=row_idx, column=col)
                        merged_cell.alignment = Alignment(horizontal='right', vertical='center')
                        
                if cell.value and str(cell.value).strip() in [
                    Setting.get_value("chief_label", "Thủ trưởng"),
                    Setting.get_value("chief_accountant_label", "Kế toán trưởng"),
                    Setting.get_value("treasurer_label", "Thủ quỹ"),
                    Setting.get_value("chief", ""),
                    Setting.get_value("chief_accountant", ""),
                    Setting.get_value("treasurer", ""),
                ]:
                    col = cell.column
                    row_idx = cell.row
                    if col < sheet.max_column:  # tránh merge vượt giới hạn
                        sheet.merge_cells(start_row=row_idx, start_column=col,
                                          end_row=row_idx, end_column=col+1)
                        merged_cell = sheet.cell(row=row_idx, column=col)
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Save dialog
        save_path, _ = QFileDialog.getSaveFileName(self, "Lưu sổ quỹ", "So_Quy.xlsx", "Excel Files (*.xlsx)")
        if save_path:
            try:
                workbook.save(save_path)
                QMessageBox.information(self, "Thành công", f"Sổ quỹ đã được lưu tại:\n{save_path}")
            except PermissionError:
                QMessageBox.warning(
                    self,
                    "Không thể lưu file",
                    "File đang được mở ở ứng dụng khác (ví dụ Excel).\n"
                    "Vui lòng đóng file và thử lại.",
                )
            except Exception as e:
                QMessageBox.critical(
                    self, "Lỗi không xác định", f"Không thể lưu file:\n{e}"
                )